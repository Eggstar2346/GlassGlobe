import requests
import xlrd
import xlwt
from openpyxl import load_workbook
import csv
import json

from flask import Flask, jsonify, render_template, request
app = Flask(__name__)



def getCoord():
    loc = ("/Users/Owner/Documents/Hack the Globe/USA Data.xlsx")
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)
    sheet.cell_value(0,0)

    print(sheet.nrows) #extract number of rows
    print(sheet.ncols) #extract number of cols
    coord = []
    for i in range (0,sheet.ncols): #put all the lat, lon coordinates into list, coord
        if (i != 0):
            coord = coord + [[sheet.cell_value(i,2),sheet.cell_value(i,3), sheet.cell_value(i,9), sheet.cell_value(i,0)]] #latitude, longitude, population, city name
    return coord
#print(coord)

coord = getCoord()

#wind speed
def getWind():
    wind = []
    #wb = load_workbook("/Users/Owner/Documents/Hack the Globe/USA Data.xlsx")
    #ws = wb.get_sheet_by_name("Sheet1")

#print(len(coord))

    for i in range (0,len(coord)):
    #change parameters
        response = requests.get("http://api.openweathermap.org/data/2.5/weather?q=" + str(coord[i][3]) + ",us&APPID=2e039e401e443365c0d9005f4cfdae59").json()
    #sheet.write(i+1,10,response['wind'])
    # c = ws.cell(row=2+i, column=12)
    # try:
    #     #c.value = (response['wind']['speed'])
    #     wind = wind + [response['wind']['speed']]
    # except:
    #     #c.value = (0)
    #     wind = wind + [0]
        wind = wind + [response['wind']['speed']]
    #print(response)
    return wind

#wb.save("/Users/Owner/Documents/Hack the Globe/USA Data.xlsx")
print(getWind())


#population
def getPop():
    pop = []
    for i in range (0,len(coord)):
        pop = pop + [coord[i][2]]

    return pop
#print(pop)

@app.route('/map')
def map(flooded):
    wind = getWind()
    hurricane = []
    nohurricane = []
    for i in wind:
        if (i > 33):
            hurricane = hurricane + [i]
        else:
            nohurricane = nohurricane + [i]
            
    if (flooded):
        return jsonify(result=hurricane)
    else:
        return jsonify(result=nohurricane)

@app.route('/')
def index():
    return render_template('index.html')
